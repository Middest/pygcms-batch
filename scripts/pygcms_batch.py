#!/usr/bin/env python3
"""
Py-GC-MS Batch Analysis Script
================================
Batch process NIST search export TXT files from Py-GC-MS analysis.
Filters SI>=80, classifies compounds by Chen 2023 and Kallenbach 2016 schemes,
computes elemental composition, NOSC, and ΔG_COX.

Usage:
    python pygcms_batch.py --input <dir> --output <output.xlsx> [--sample_map <json>]
"""

import os
import re
import json
import argparse
import time
import urllib.request
import urllib.parse
import urllib.error
import ssl
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Path to the built-in lookup table (relative to this script)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_LOOKUP = os.path.join(_SCRIPT_DIR, '..', 'data', 'chen_lookup.json')


# ============================================================================
# 1. NIST TXT PARSER
# ============================================================================

def parse_molecular_formula(formula_str):
    """
    Parse a molecular formula string into atom counts.
    Returns dict with C, H, O, N, P, S, other counts.
    Handles formulas like 'C6H6O3', 'C5H14N2', 'C11H23Cl', etc.
    """
    atoms = {'C': 0, 'H': 0, 'O': 0, 'N': 0, 'P': 0, 'S': 0, 'other': {}}

    if not formula_str or formula_str == '?' or formula_str == '':
        return atoms

    # Remove any whitespace
    formula_str = formula_str.strip()

    # Pattern: element symbol followed by optional count
    pattern = re.compile(r'([A-Z][a-z]?)(\d*)')
    matches = pattern.findall(formula_str)

    for elem, count in matches:
        c = int(count) if count else 1
        if elem in atoms:
            atoms[elem] = c
        else:
            atoms['other'][elem] = c

    return atoms


def compute_nosc_and_dg(atoms):
    """
    Compute NOSC (Nominal Oxidation State of Carbon) and ΔG_COX.

    NOSC = -((-Z + 4*C + H - 3*N - 2*O + 5*P - 2*S) / C)
    where Z = net charge (assumed 0 for neutral molecules)

    ΔG_COX = 60.3 - 28.5 * NOSC  (kJ per mol C)

    Reference: LaRowe & Van Cappellen 2011; Wu et al. 2025 Nature Food
    """
    C = atoms.get('C', 0)
    H = atoms.get('H', 0)
    N = atoms.get('N', 0)
    O = atoms.get('O', 0)
    P = atoms.get('P', 0)
    S = atoms.get('S', 0)
    Z = 0  # neutral molecules

    if C == 0:
        return None, None

    nosc = -((-Z + 4*C + H - 3*N - 2*O + 5*P - 2*S) / C)
    dg_cox = 60.3 - 28.5 * nosc

    return round(nosc, 6), round(dg_cox, 3)


def compute_hc_oc(atoms):
    """Compute H/C and O/C atomic ratios from atom counts."""
    C = atoms.get('C', 0)
    H = atoms.get('H', 0)
    O = atoms.get('O', 0)
    if C == 0:
        return None, None
    return round(H / C, 4), round(O / C, 4)


# ============================================================================
# 1.5 NIST Chemistry WebBook Online Query
# ============================================================================

# Cache to avoid repeated queries
_nist_webbook_cache = {}
_nist_cache_file = None


def _get_cache_path():
    """Get path to NIST WebBook cache file."""
    return os.path.join(_SCRIPT_DIR, '..', 'data', 'nist_webbook_cache.json')


def _load_cache():
    """Load NIST WebBook cache from disk."""
    global _nist_webbook_cache, _nist_cache_file
    _nist_cache_file = _get_cache_path()
    if os.path.exists(_nist_cache_file):
        try:
            with open(_nist_cache_file, 'r', encoding='utf-8') as f:
                _nist_webbook_cache = json.load(f)
        except Exception:
            _nist_webbook_cache = {}
    return _nist_webbook_cache


def _save_cache():
    """Save NIST WebBook cache to disk."""
    if _nist_cache_file:
        os.makedirs(os.path.dirname(_nist_cache_file), exist_ok=True)
        with open(_nist_cache_file, 'w', encoding='utf-8') as f:
            json.dump(_nist_webbook_cache, f, ensure_ascii=False, indent=2)


def query_nist_webbook_by_cas(cas, use_cache=True):
    """
    Query NIST Chemistry WebBook by CAS number.
    Returns dict with: formula, mol_weight, inchi, inchikey, iupac_name, cas_verified
    Returns None if not found.
    """
    if not cas or cas in ('0 - 00 - 0', '0-00-0', 'N/A', ''):
        return None

    cache = _load_cache() if use_cache else {}
    cache_key = f'CAS:{cas}'
    if cache_key in cache:
        return cache[cache_key]

    # Clean CAS: remove spaces, keep digits and hyphens
    cas_clean = re.sub(r'\s+', '', cas)
    cas_digits = cas_clean.replace('-', '')

    url = f'https://webbook.nist.gov/cgi/cbook.cgi?ID=C{cas_digits}&Units=SI&Mask=2000'

    try:
        # Allow self-signed certs for resilience
        ctx = ssl.create_default_context()
        req = urllib.request.Request(url, headers={'User-Agent': 'pygcms-batch/1.0 (academic)'})
        with urllib.request.urlopen(req, timeout=20, context=ctx) as resp:
            html = resp.read().decode('utf-8', errors='ignore')

        result = {}

        # Extract JSON-LD structured data
        formula_m = re.search(r'"molecularFormula"\s*:\s*"([^"]+)"', html)
        if formula_m:
            result['formula'] = formula_m.group(1)

        mw_m = re.search(r'"molecularWeight"\s*:\s*"([^"]+)"', html)
        if mw_m:
            result['mol_weight'] = mw_m.group(1)

        inchi_m = re.search(r'"inChI"\s*:\s*"([^"]+)"', html)
        if inchi_m:
            result['inchi'] = inchi_m.group(1)

        inchikey_m = re.search(r'"inChIKey"\s*:\s*"([^"]+)"', html)
        if inchikey_m:
            result['inchikey'] = inchikey_m.group(1)

        name_m = re.search(r'"name"\s*:\s*"([^"]+)"', html)
        if name_m:
            result['iupac_name'] = name_m.group(1)

        # Check if CAS was found
        if 'Species not found' in html or 'no data' in html.lower():
            result['status'] = 'not_found'
            cache[cache_key] = None
            _save_cache()
            return None

        if result:
            result['status'] = 'found'
            result['cas_queried'] = cas_clean
            cache[cache_key] = result
            _save_cache()
            return result
        else:
            cache[cache_key] = None
            _save_cache()
            return None

    except Exception as e:
        # Network error or timeout — don't cache, allow retry
        return {'status': 'error', 'error': str(e)}


def query_nist_webbook_by_name(name, use_cache=True):
    """
    Query NIST Chemistry WebBook by compound name.
    Slower and less reliable than CAS lookup. Use CAS when available.
    """
    if not name or len(name) < 3:
        return None

    cache = _load_cache() if use_cache else {}
    cache_key = f'NAME:{name[:80]}'
    if cache_key in cache:
        return cache[cache_key]

    # URL-encode the name
    encoded_name = urllib.parse.quote(name[:80])
    url = f'https://webbook.nist.gov/cgi/cbook.cgi?Name={encoded_name}&Units=SI&Mask=2000'

    try:
        ctx = ssl.create_default_context()
        req = urllib.request.Request(url, headers={'User-Agent': 'pygcms-batch/1.0 (academic)'})
        with urllib.request.urlopen(req, timeout=20, context=ctx) as resp:
            html = resp.read().decode('utf-8', errors='ignore')

        if 'no data' in html.lower() or 'Species not found' in html:
            cache[cache_key] = None
            _save_cache()
            return None

        result = {}
        formula_m = re.search(r'"molecularFormula"\s*:\s*"([^"]+)"', html)
        if formula_m:
            result['formula'] = formula_m.group(1)
        mw_m = re.search(r'"molecularWeight"\s*:\s*"([^"]+)"', html)
        if mw_m:
            result['mol_weight'] = mw_m.group(1)
        inchi_m = re.search(r'"inChI"\s*:\s*"([^"]+)"', html)
        if inchi_m:
            result['inchi'] = inchi_m.group(1)
        inchikey_m = re.search(r'"inChIKey"\s*:\s*"([^"]+)"', html)
        if inchikey_m:
            result['inchikey'] = inchikey_m.group(1)

        if result:
            result['status'] = 'found'
            cache[cache_key] = result
            _save_cache()
            return result
        else:
            cache[cache_key] = None
            _save_cache()
            return None

    except Exception as e:
        return {'status': 'error', 'error': str(e)}


def enrich_compound_with_webbook(name, cas, existing_formula=''):
    """
    Enrich a compound entry with NIST WebBook data.
    Prioritizes CAS lookup, falls back to name lookup.
    Returns dict with any new fields found, or empty dict.
    """
    result = {}

    # Only query if we're missing key data
    need_formula = not existing_formula or len(existing_formula) < 2

    # Try CAS first (most reliable)
    wb_data = query_nist_webbook_by_cas(cas)
    if wb_data and wb_data.get('status') == 'found':
        if need_formula and wb_data.get('formula'):
            result['formula'] = wb_data['formula']
        if wb_data.get('inchi'):
            result['inchi'] = wb_data['inchi']
        if wb_data.get('inchikey'):
            result['inchikey'] = wb_data['inchikey']
        if wb_data.get('mol_weight'):
            result['mol_weight_wb'] = wb_data['mol_weight']
        return result

    # Fall back to name lookup if CAS failed
    if need_formula:
        wb_data = query_nist_webbook_by_name(name)
        if wb_data and wb_data.get('status') == 'found':
            if need_formula and wb_data.get('formula'):
                result['formula'] = wb_data['formula']
            if wb_data.get('inchi'):
                result['inchi'] = wb_data['inchi']
            if wb_data.get('inchikey'):
                result['inchikey'] = wb_data['inchikey']
            return result

    return result

def parse_nist_txt(filepath):
    """
    Parse a NIST search export TXT file.
    Extracts data from BOTH [MC Peak Table] (for TIC area) and
    [MS Similarity Search Results for Spectrum Process Table] (for SI, CAS, formula).

    Returns list of dicts, each containing merged peak data for Hit#1 entries with SI>=80.
    """
    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    # === Step 1: Parse MS Similarity Search Results (Hit #1 only) ===
    # This is the authoritative source for SI, CAS, molecular formula
    nist_data = {}  # spectrum_num -> {si, cas, name, formula, mol_weight}

    sim_match = re.search(
        r'\[MS Similarity Search Results for Spectrum Process Table\](.*?)(?=\n\[MS Chromatogram\])',
        content, re.DOTALL
    )
    if sim_match:
        sim_text = sim_match.group(1)
        lines = sim_text.strip().split('\n')
        in_data = False
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith('Spectrum#'):
                in_data = True
                continue
            if in_data and line.startswith('['):
                break
            if in_data:
                parts = line.split('\t')
                if len(parts) >= 7:
                    try:
                        spectrum_num = int(parts[0])
                        hit_num = int(parts[1])
                        si = int(parts[2]) if parts[2] else 0
                        cas = parts[3].strip() if len(parts) > 3 else ''
                        name = parts[4].strip() if len(parts) > 4 else ''
                        mol_weight = parts[5].strip() if len(parts) > 5 else ''
                        formula = parts[6].strip() if len(parts) > 6 else ''

                        # Only keep Hit #1 (the best match) for each spectrum
                        if hit_num == 1 and spectrum_num not in nist_data:
                            # Clean name: remove $$ and trailing content
                            clean_name = name.split('$$')[0].strip()
                            nist_data[spectrum_num] = {
                                'si': si,
                                'cas': cas,
                                'name': clean_name,
                                'formula': formula,
                                'mol_weight': mol_weight
                            }
                    except (ValueError, IndexError):
                        pass

    # === Step 2: Parse MC Peak Table for TIC areas ===
    # Match with NIST data by peak number (Peak# in MC = Spectrum# in MS results)
    mc_peaks = {}  # peak_num -> area, ret_time

    mc_match = re.search(r'\[MC Peak Table\](.*?)(?=\n\[)', content, re.DOTALL)
    if not mc_match:
        mc_match = re.search(r'\[MC Peak Table\](.*)', content, re.DOTALL)
    if mc_match:
        mc_text = mc_match.group(1)
        lines = mc_text.strip().split('\n')
        in_data = False
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if line.startswith('Peak#'):
                in_data = True
                continue
            if in_data and line.startswith('['):
                break
            if in_data:
                parts = line.split('\t')
                if len(parts) >= 7:
                    try:
                        peak_num = int(parts[0])
                        ret_time = float(parts[1])
                        # Area is at index 5 (0-based) = 6th column
                        area = float(parts[5]) if parts[5] else 0
                        mc_peaks[peak_num] = {
                            'ret_time': ret_time,
                            'area': area
                        }
                    except (ValueError, IndexError):
                        pass

    # === Step 3: Merge NIST data with MC Peak Table data ===
    # Build ALL peaks (including SI<80) and SI>=80 filtered peaks
    all_peaks = []
    filtered_peaks = []

    for spectrum_num, nist in nist_data.items():
        mc = mc_peaks.get(spectrum_num, {})
        area = mc.get('area', 0)
        ret_time = mc.get('ret_time', 0)

        # Parse molecular formula
        atoms = {}
        if nist['formula']:
            atoms = parse_molecular_formula(nist['formula'])
        nosc, dg = compute_nosc_and_dg(atoms)
        hc, oc = compute_hc_oc(atoms)

        peak_entry = {
            'peak_num': spectrum_num,
            'ret_time': ret_time,
            'area': area,
            'name': nist['name'],
            'si': nist['si'],
            'cas': nist['cas'],
            'formula': nist['formula'],
            'mol_weight': nist['mol_weight'],
            'atoms': atoms,
            'nosc': nosc,
            'dg_cox': dg,
            'hc': hc,
            'oc': oc
        }

        all_peaks.append(peak_entry)
        if nist['si'] >= 80:
            filtered_peaks.append(peak_entry)

    # Also include MC peaks that had no NIST hit (SI=0, no formula)
    for peak_num, mc in mc_peaks.items():
        if peak_num not in nist_data:
            all_peaks.append({
                'peak_num': peak_num,
                'ret_time': mc.get('ret_time', 0),
                'area': mc.get('area', 0),
                'name': '',
                'si': 0,
                'cas': '',
                'formula': '',
                'mol_weight': '',
                'atoms': {'C': 0, 'H': 0, 'O': 0, 'N': 0, 'P': 0, 'S': 0, 'other': {}},
                'nosc': None,
                'dg_cox': None
            })

    return all_peaks, filtered_peaks


# ============================================================================
# 2. COMPOUND CLASSIFICATION
# ============================================================================

# --- Chen 2023 Classification Rules ---
# Chemical groups and source attribution based on:
# Chen et al. 2023, Carbon Research 2:1

CHEN_CLASS_RULES = {
    'lipids': {
        'keywords': ['acid', 'ester', 'alkane', 'alkene', 'alcohol', 'ketone',
                     'aldehyde', 'fatty', 'glycerol', 'sterol', 'terpene',
                     'acetate', 'propanoate', 'butanoate', 'hexanoate',
                     'tridecanoic', 'tetradecanoic', 'pentadecanoic',
                     'hexadecanoic', 'heptadecanoic', 'octadecanoic',
                     'eicosanoic', 'docosanoic', 'heneicosanoic',
                     'decane', 'undecane', 'dodecane', 'tridecane',
                     'tetradecane', 'pentadecane', 'hexadecane',
                     'heptadecane', 'octadecane', 'nonadecane', 'eicosane',
                     'heneicosane', 'docosane', 'tricosane', 'tetracosane',
                     'decene', 'undecene', 'dodecene', 'tridecene',
                     'tetradecene', 'pentadecene', 'hexadecene',
                     'heptadecene', 'octadecene', 'nonadecene', 'eicosene',
                     'sterol', 'cholesterol', 'campesterol', 'sitosterol',
                     'stigmasterol', 'ergosterol',
                     'alkyne', 'squalene', 'tocopherol', 'carotene',
                     'decanoic', 'dodecanoic', 'octadecanenitrile',
                     'heptadecanenitrile', 'hexadecanenitrile',
                     'octanol', 'decanol', 'dodecanol', 'tetradecanol',
                     'hexadecanol', 'octadecanol'],
        'source': 'mixed'  # short-chain (<16C) microbial, long-chain (>18C) plant, 16-18C mixed
    },
    'monocyclic_aromatics': {
        'keywords': ['benzene', 'toluene', 'xylene', 'ethylbenzene', 'styrene',
                     'phenol', 'cresol', 'benzaldehyde', 'benzoic',
                     'benzyl', 'phenyl', 'anisole', 'catechol', 'guaiacol',
                     'syringol', 'trimethylbenzene', 'propylbenzene',
                     'butylbenzene', 'isopropylbenzene', 'cumene',
                     'indane', 'indene', 'tetralin',
                     'acetophenone', 'benzophenone',
                     'methylphenol', 'dimethylphenol', 'ethylphenol',
                     'methoxybenzene', 'dimethoxybenzene',
                     'trimethoxybenzene', 'hydroxybenzene'],
        'source': 'microbial'
    },
    'polycyclic_aromatics': {
        'keywords': ['naphthalene', 'anthracene', 'phenanthrene', 'pyrene',
                     'fluoranthene', 'chrysene', 'benzopyrene', 'fluorene',
                     'biphenyl', 'indole', 'quinoline', 'isoquinoline',
                     'carbazole', 'dibenzofuran', 'dibenzothiophene',
                     'acenaphthene', 'acenaphthylene', 'benzo[a]',
                     'methylnaphthalene', 'dimethylnaphthalene',
                     'trimethylnaphthalene', 'tetramethylnaphthalene',
                     'retene', 'perylene', 'coronene', 'benzoperylene',
                     'indenopyrene', 'benzofluoranthene', 'benzanthracene'],
        'source': 'microbial'
    },
    'phenolics': {
        'keywords': ['phenol', 'cresol', 'catechol', 'resorcinol',
                     'hydroquinone', 'guaiacol', 'syringol', 'vanillin',
                     'acetovanillone', 'acetosyringone', 'coniferyl',
                     'sinapyl', 'coumaryl', 'ferulic', 'p-coumaric',
                     'cinnamic', 'benzoic acid', 'salicylic',
                     'protocatechuic', 'gallic', 'syringic', 'vanillic',
                     'eugenol', 'isoeugenol', 'chavicol',
                     'dihydroxybenzene', 'trihydroxybenzene',
                     'methoxyphenol', 'dimethoxyphenol',
                     'hydroxybenzoic', 'dihydroxybenzoic',
                     'hydroxyphenyl', 'dihydroxyphenyl',
                     'methoxybenzoic', 'dimethoxybenzoic',
                     'hydroxybenzaldehyde', 'hydroxyacetophenone',
                     'phenylpropanoid', 'flavonoid', 'tannin',
                     'lignin monomer', 'monolignol'],
        'source': 'plant'
    },
    'polysaccharides': {
        'keywords': ['furan', 'furfural', 'furanone', 'pyran', 'pyranone',
                     'levoglucosan', 'levoglucosenone', 'mannosan',
                     'galactosan', 'xylosan', 'arabinosan',
                     'anhydro sugar', 'anhydrosugar', 'cyclohexanone',
                     'furoic', 'furan methanol', 'furanmethanol',
                     'furfuryl', 'methylfuran', 'dimethylfuran',
                     'hydroxymethylfurfural', 'acetic acid',
                     'propanoic acid', 'butanoic acid',
                     'pentanoic acid', 'hexanoic acid',
                     'dianhydromannitol', 'dianhydrosorbitol',
                     'dianhydroglucitol', 'glucose', 'mannose',
                     'galactose', 'xylose', 'arabinose', 'ribose',
                     'sugar', 'glycoside', 'cellulose marker',
                     'hemicellulose marker', 'chitin marker',
                     'acetylfuran', 'methylfuroate',
                     'acetamide, n-methyl', 'acetamide, n-ethyl',
                     'polysaccharide-derived', 'carbohydrate-derived',
                     'acetic acid, methyl ester', 'acetic acid, ethyl ester',
                     '2-cyclopenten-1-one', 'methyl cyclopentenone'],
        'source': 'plant'
    },
    'lignins': {
        'keywords': ['lignin', 'guaiacyl', 'syringyl', 'hydroxyphenyl',
                     'coniferyl alcohol', 'sinapyl alcohol', 'coumaryl alcohol',
                     'dihydroconiferyl alcohol', 'propylguaiacol',
                     'vinylguaiacol', 'vinylphenol', 'vinylsyringol',
                     'propenylguaiacol', 'propenylsyringol',
                     'allylguaiacol', 'allylsyringol',
                     'ethylguaiacol', 'ethylsyringol', 'methylguaiacol',
                     'homovanillic', 'homosyringic',
                     'guaiacylethane', 'syringylethane',
                     'guaiacylpropane', 'syringylpropane',
                     'guaiacylpropanol', 'syringylpropanol',
                     'guaiacylpropanone', 'syringylpropanone',
                     'guaiacylacetic acid', 'syringylacetic acid',
                     'lignin-derived phenol', 'lignin marker',
                     'methoxyeugenol', 'dimethoxyeugenol',
                     'trans-isoeugenol', 'trans-propenylsyringol',
                     '4-vinylguaiacol', '4-vinylsyringol',
                     '4-methylguaiacol', '4-ethylguaiacol',
                     '4-methylsyringol', '4-ethylsyringol',
                     'coniferaldehyde', 'sinapaldehyde'],
        'source': 'plant'
    },
    'amino_N_bearing': {
        'keywords': ['nitrile', 'pyridine', 'pyrrole', 'pyrazine', 'imidazole',
                     'indole', 'amine', 'amino', 'amide',
                     'piperidine', 'piperazine', 'pyrrolidine',
                     'pyrimidine', 'purine', 'adenine', 'guanine',
                     'methylpyridine', 'dimethylpyridine', 'trimethylpyridine',
                     'methylpyrrole', 'dimethylpyrrole',
                     'methylimidazole', 'dimethylimidazole',
                     'methylpyrazine', 'dimethylpyrazine', 'trimethylpyrazine',
                     'butanenitrile', 'pentanenitrile', 'hexanenitrile',
                     'heptanenitrile', 'octanenitrile', 'decanenitrile',
                     'propanenitrile', 'benzonitrile', 'acetonitrile',
                     'methylbutanenitrile', 'dimethylbutanenitrile',
                     'aminobenzoic', 'aminophenol',
                     'protein marker', 'amino acid derivative',
                     'diketopiperazine', 'diketodipyrrole',
                     'proline', 'tryptophan', 'tyrosine', 'phenylalanine',
                     'valine', 'leucine', 'isoleucine', 'alanine', 'glycine'],
        'source': 'mixed'  # nitriles/pyridines/pyrroles= microbial; indoles = plant
    },
    'heterocyclic_N_bearing': {
        'keywords': ['pyrazole', 'oxazole', 'thiazole', 'triazole',
                     'tetrazole', 'diazine', 'triazine',
                     'benzimidazole', 'benzoxazole', 'benzothiazole',
                     'benzotriazole', 'quinoxaline', 'quinazoline',
                     'purine base', 'pyrimidine base',
                     'uracil', 'thymine', 'cytosine',
                     'xanthine', 'hypoxanthine', 'theobromine', 'caffeine',
                     'riboflavin', 'niacin', 'nicotinamide',
                     'nucleobase', 'nucleoside',
                     'pteridine', 'pterin', 'flavin', 'folate',
                     'adenine derivative', 'guanine derivative'],
        'source': 'plant'  # mainly plant-derived (pyrazines)
    },
    'other_N_bearing': {
        'keywords': ['nitro', 'nitroso', 'hydroxylamine', 'hydrazine',
                     'azo', 'diazo', 'azide', 'cyanate', 'isocyanate',
                     'thiocyanate', 'isothiocyanate',
                     'nitrile oxide', 'oxime', 'oxazole',
                     'ammonium', 'quaternary ammonium',
                     'betaine', 'choline', 'carnitine',
                     'urea', 'thiourea', 'guanidine', 'creatine',
                     'methylamine', 'dimethylamine', 'trimethylamine',
                     'ethylamine', 'diethylamine', 'triethylamine',
                     'propylamine', 'butylamine',
                     'putrescine', 'cadaverine', 'spermidine', 'spermine',
                     'ethanolamine', 'diethanolamine', 'triethanolamine',
                     'methanediamine', 'ethanediamine',
                     'tetramethylammonium', 'tetraethylammonium',
                     'n,n-dimethyl', 'n,n-diethyl', 'n,n,n\',n\'-tetramethyl',
                     'dimethylamino', 'diethylamino',
                     'n-bearing', 'n-containing', 'nitrogen-containing'],
        'source': 'microbial'
    },
    'unspecified': {
        'keywords': ['chloro', 'bromo', 'fluoro', 'iodo', 'chloride',
                     'bromide', 'fluoride', 'iodide',
                     'silane', 'siloxane', 'silicone',
                     'phosphate', 'phosphonate', 'phosphine',
                     'sulfone', 'sulfoxide', 'sulfonate', 'sulfate',
                     'sulfur', 'sulfide', 'disulfide', 'thiol', 'thioether',
                     'borane', 'borate', 'silicon', 'germanium', 'arsenic',
                     'selenium', 'tellurium', 'tin',
                     'unknown', 'unidentified', 'not classified',
                     'artifact', 'column bleed', 'plasticizer',
                     'phthalate', 'adipate', 'sebacate',
                     'biphenyl, chloro', 'dichloro', 'trichloro',
                     'tetrachloro', 'pentachloro', 'hexachloro',
                     'silane,', 'siloxane,', 'cyclosiloxane',
                     'polydimethylsiloxane', 'pdms',
                     'perchlorate', 'halogen'],
        'source': 'mixed'
    }
}

# --- Kallenbach 2016 Classification Rules ---
# Based on Kallenbach et al. 2016, Nature Communications 7:13630
# Classifies compounds by origin-based categories

KALLENBACH_CLASS_RULES = {
    'lipids': {
        'keywords': ['acid', 'ester', 'alkane', 'alkene', 'alcohol', 'ketone',
                     'aldehyde', 'fatty', 'glycerol', 'sterol', 'terpene',
                     'acetate', 'propanoate', 'butanoate', 'hexanoate',
                     'decane', 'undecane', 'dodecane', 'tridecane',
                     'tetradecane', 'pentadecane', 'hexadecane',
                     'heptadecane', 'octadecane', 'nonadecane', 'eicosane',
                     'heneicosane', 'docosane', 'tricosane', 'tetracosane',
                     'decenoic', 'dodecanoic', 'tetradecanoic',
                     'hexadecanoic', 'octadecanoic',
                     'octadecanenitrile', 'heptadecanenitrile',
                     'hexadecanenitrile', 'tetradecanenitrile',
                     'octadecanamide', 'hexadecanamide', 'tetradecanamide',
                     'squalene', 'sterol', 'cholesterol', 'tocopherol',
                     'carotene', 'long-chain', 'wax ester',
                     'tridecanoic', 'tetradecanoic', 'pentadecanoic',
                     'hexadecanoic', 'heptadecanoic', 'octadecanoic',
                     'eicosanoic', 'docosanoic'],
        'source': 'mixed'
    },
    'lignin_derivatives': {
        'keywords': ['lignin', 'guaiacyl', 'syringyl', 'coniferyl', 'sinapyl',
                     'coumaryl', 'guaiacol', 'syringol', 'vanillin',
                     'acetovanillone', 'acetosyringone',
                     'vinylguaiacol', 'vinylsyringol', 'vinylphenol',
                     'ethylguaiacol', 'ethylsyringol',
                     'propylguaiacol', 'propenylsyringol',
                     'guaiacylethane', 'syringylethane',
                     'guaiacylpropane', 'syringylpropane',
                     'lignin-derived phenol', 'lignin marker',
                     'coniferaldehyde', 'sinapaldehyde',
                     'homovanillic', 'homosyringic',
                     'guaiacylpropanone', 'syringylpropanone',
                     'trans-isoeugenol', 'trans-propenylsyringol',
                     '4-vinylguaiacol', '4-vinylsyringol',
                     'methoxyeugenol', 'dimethoxyeugenol',
                     '4-methylguaiacol', '4-ethylguaiacol',
                     '4-methylsyringol', '4-ethylsyringol'],
        'source': 'plant'
    },
    'polysaccharides': {
        'keywords': ['furan', 'furfural', 'furanone', 'pyran', 'pyranone',
                     'levoglucosan', 'levoglucosenone', 'mannosan',
                     'galactosan', 'xylosan', 'arabinosan',
                     'anhydro sugar', 'anhydrosugar', 'cyclohexanone',
                     'furoic', 'furan methanol', 'furfuryl',
                     'methylfuran', 'dimethylfuran',
                     'hydroxymethylfurfural', 'cellulose marker',
                     'hemicellulose marker', 'chitin marker',
                     'acetylfuran', 'methylfuroate',
                     'glucose', 'mannose', 'galactose',
                     'xylose', 'arabinose', 'ribose',
                     'sugar', 'glycoside', 'polysaccharide-derived',
                     'dianhydromannitol', 'dianhydrosorbitol',
                     'dianhydroglucitol'],
        'source': 'mixed'  # plant polysaccharides + microbial (chitin)
    },
    'proteins': {
        'keywords': ['nitrile', 'pyridine', 'pyrrole', 'pyrazine', 'imidazole',
                     'indole', 'amine', 'amino', 'amide',
                     'piperidine', 'piperazine', 'pyrrolidine',
                     'pyrimidine', 'purine',
                     'diketopiperazine', 'diketodipyrrole',
                     'proline', 'tryptophan', 'tyrosine', 'phenylalanine',
                     'valine', 'leucine', 'isoleucine', 'alanine', 'glycine',
                     'protein marker', 'amino acid derivative',
                     'acetonitrile', 'propanenitrile', 'butanenitrile',
                     'pentanenitrile', 'hexanenitrile', 'benzonitrile',
                     'methylpyridine', 'dimethylpyridine', 'trimethylpyridine',
                     'methylpyrrole', 'dimethylpyrrole',
                     'methylimidazole', 'dimethylimidazole',
                     'methylpyrazine', 'dimethylpyrazine', 'trimethylpyrazine',
                     'methanediamine', 'ethanediamine',
                     'n,n-dimethyl', 'n,n-diethyl', 'tetramethyl',
                     'dimethylamino', 'diethylamino'],
        'source': 'microbial'
    },
    'non_protein_N': {
        'keywords': ['nitro', 'nitroso', 'hydroxylamine', 'hydrazine',
                     'azo', 'diazo', 'cyanate', 'isocyanate',
                     'thiocyanate', 'isothiocyanate',
                     'ammonium', 'quaternary ammonium', 'betaine',
                     'choline', 'carnitine', 'urea', 'thiourea',
                     'guanidine', 'creatine', 'creatinine',
                     'tetramethylammonium', 'tetraethylammonium',
                     'adenine', 'guanine', 'uracil', 'thymine', 'cytosine',
                     'xanthine', 'hypoxanthine', 'theobromine', 'caffeine',
                     'nucleobase', 'nucleoside', 'riboflavin', 'niacin',
                     'purine base', 'pyrimidine base',
                     'pteridine', 'pterin', 'flavin', 'folate'],
        'source': 'mixed'
    },
    'phenolics': {
        'keywords': ['phenol', 'cresol', 'catechol', 'resorcinol',
                     'hydroquinone', 'vanillin', 'vanillic',
                     'syringic', 'gallic', 'protocatechuic',
                     'p-coumaric', 'ferulic', 'cinnamic',
                     'benzoic acid', 'salicylic',
                     'hydroxybenzoic', 'dihydroxybenzoic',
                     'hydroxyphenyl', 'dihydroxyphenyl',
                     'methoxyphenol', 'dimethoxyphenol',
                     'eugenol', 'isoeugenol', 'chavicol',
                     'hydroxybenzaldehyde', 'hydroxyacetophenone',
                     'acetovanillone', 'acetosyringone',
                     'dihydroxybenzene', 'trihydroxybenzene',
                     'phenylpropanoid'],
        'source': 'plant'
    },
    'aromatics': {
        'keywords': ['benzene', 'toluene', 'xylene', 'ethylbenzene', 'styrene',
                     'naphthalene', 'anthracene', 'phenanthrene', 'pyrene',
                     'fluoranthene', 'fluorene', 'biphenyl',
                     'benzyl', 'phenyl', 'benzaldehyde',
                     'trimethylbenzene', 'propylbenzene',
                     'butylbenzene', 'isopropylbenzene', 'cumene',
                     'indane', 'indene', 'tetralin',
                     'acetophenone', 'benzophenone',
                     'methylnaphthalene', 'dimethylnaphthalene',
                     'methoxybenzene', 'dimethoxybenzene',
                     'trimethoxybenzene',
                     'acenaphthene', 'acenaphthylene',
                     'retene', 'perylene', 'coronene',
                     'benzoperylene', 'chrysene',
                     'methylphenol', 'dimethylphenol', 'ethylphenol',
                     'anisole', 'benzonitrile'],
        'source': 'microbial'
    },
    'unspecified': {
        'keywords': ['chloro', 'bromo', 'fluoro', 'iodo', 'chloride',
                     'bromide', 'fluoride', 'iodide',
                     'silane', 'siloxane', 'silicone',
                     'phosphate', 'phosphonate', 'phosphine',
                     'sulfone', 'sulfoxide', 'sulfonate', 'sulfate',
                     'sulfur', 'sulfide', 'disulfide', 'thiol',
                     'borane', 'borate', 'silicon',
                     'unknown', 'unidentified', 'not classified',
                     'artifact', 'column bleed', 'plasticizer',
                     'phthalate', 'adipate', 'sebacate',
                     'perchlorate', 'halogen',
                     'polydimethylsiloxane', 'pdms', 'cyclosiloxane',
                     'benzyl oxy', 'digitoxin'],
        'source': 'mixed'
    }
}


# Global lookup cache (loaded at first use)
_chen_lookup = None


def _load_lookup(lookup_path=None):
    """Load Chen classification lookup table from JSON."""
    global _chen_lookup
    if _chen_lookup is not None:
        return _chen_lookup

    if lookup_path is None:
        lookup_path = _DEFAULT_LOOKUP

    if os.path.exists(lookup_path):
        with open(lookup_path, 'r', encoding='utf-8') as f:
            _chen_lookup = json.load(f)
    else:
        _chen_lookup = {}

    return _chen_lookup


def _normalize_name(name):
    """Normalize compound name for lookup matching."""
    # Remove trailing/leading whitespace, collapse internal spaces
    name = ' '.join(name.split())
    # Remove common NIST suffixes ($$, etc.)
    name = re.sub(r'\s*\$\$.*$', '', name)
    return name.strip()


def _lookup_match(name, lookup):
    """Try to match a compound name against the lookup table."""
    name_clean = _normalize_name(name)

    # 1. Exact match
    if name_clean in lookup:
        return lookup[name_clean]

    # 2. Case-insensitive match
    name_lower = name_clean.lower()
    for k, v in lookup.items():
        if k.lower() == name_lower:
            return v

    # 3. Partial match (first 40 chars match, handles NIST name variants)
    name_prefix = name_lower[:40]
    for k, v in lookup.items():
        if k.lower()[:40] == name_prefix:
            return v

    # 4. Substring match (name contains lookup key or vice versa)
    for k, v in lookup.items():
        kl = k.lower()
        if len(kl) > 10 and (kl in name_lower or name_lower in kl):
            return v

    return None


def classify_compound_chen(name, lookup=None):
    """
    Classify a compound by Chen 2023 scheme.
    Uses exact lookup table first, falls back to keyword matching.
    Returns (chemical_group, source_category)
    """
    if lookup is None:
        lookup = _load_lookup()

    # Step 1: Try exact lookup
    match = _lookup_match(name, lookup)
    if match:
        chen_class = match.get('class', '')
        chen_source = match.get('source', 'mixed')
        # Map POC-style class names to internal keys
        class_map = {
            'Lipids': 'lipids',
            'Other N-bearing': 'other_N_bearing',
            'Amino N-bearing (nitrile/pyridine/pyrrole)': 'amino_N_bearing',
            'Monocyclic aromatics': 'monocyclic_aromatics',
            'Unspecified/halogen_or_complex': 'unspecified',
            'Polysaccharides / carbohydrate-derived': 'polysaccharides',
            'Polycyclic aromatics': 'polycyclic_aromatics',
            'Phenolics': 'phenolics',
            'Unspecified/complex': 'unspecified',
            'Lignin derivatives': 'lignins',
        }
        chen_group = class_map.get(chen_class, 'unspecified')
        return (chen_group, chen_source)

    # Step 2: Fall back to keyword matching
    name_lower = name.lower().strip()
    for group, rules in CHEN_CLASS_RULES.items():
        for kw in rules['keywords']:
            if kw.lower() in name_lower:
                return (group, rules['source'])

    return ('unspecified', 'mixed')


def classify_compound_kallenbach(name, lookup=None):
    """
    Classify a compound by Kallenbach 2016 scheme.
    Uses exact lookup table first, falls back to keyword matching.
    Returns (chemical_group, source_category)
    """
    if lookup is None:
        lookup = _load_lookup()

    # Step 1: Try exact lookup
    match = _lookup_match(name, lookup)
    if match:
        chen_class = match.get('class', '')
        chen_source = match.get('source', 'mixed')
        # Map Chen classes to Kallenbach classes
        kal_map = {
            'Lipids': ('lipids', chen_source),
            'Other N-bearing': ('non_protein_N', chen_source),
            'Amino N-bearing (nitrile/pyridine/pyrrole)': ('proteins', chen_source),
            'Monocyclic aromatics': ('aromatics', chen_source),
            'Polycyclic aromatics': ('aromatics', chen_source),
            'Unspecified/halogen_or_complex': ('unspecified', 'mixed'),
            'Polysaccharides / carbohydrate-derived': ('polysaccharides', chen_source),
            'Phenolics': ('phenolics', chen_source),
            'Unspecified/complex': ('unspecified', 'mixed'),
            'Lignin derivatives': ('lignin_derivatives', chen_source),
        }
        if chen_class in kal_map:
            return kal_map[chen_class]
        return ('unspecified', 'mixed')

    # Step 2: Fall back to keyword matching
    name_lower = name.lower().strip()
    for group, rules in KALLENBACH_CLASS_RULES.items():
        for kw in rules['keywords']:
            if kw.lower() in name_lower:
                return (group, rules['source'])

    return ('unspecified', 'mixed')


# ============================================================================
# 3. BATCH PROCESSING
# ============================================================================

def process_directory(input_dir, sample_map=None, enrich=False, enrich_delay=0.5):
    """
    Process all TXT files in directory recursively.

    Args:
        input_dir: path to directory containing TXT files
        sample_map: dict mapping sample code to treatment name, e.g. {'5':'CK'}
        enrich: if True, query NIST WebBook for missing formulas
        enrich_delay: seconds between WebBook queries (be polite)

    Returns:
        dict with sample-level results
    """
    if sample_map is None:
        sample_map = {}

    results = {}
    total_enriched = 0

    for root, dirs, files in os.walk(input_dir):
        for fname in files:
            if not fname.lower().endswith('.txt'):
                continue

            filepath = os.path.join(root, fname)

            # Determine sample code from filename (strip extension)
            sample_code = os.path.splitext(fname)[0]
            treatment = sample_map.get(sample_code, sample_code)

            # Check if this is a retest (in a *复测* directory)
            subdir_name = os.path.basename(root)
            if '复测' in subdir_name or 'retest' in subdir_name.lower():
                treatment = f"{treatment}_复测"

            print(f"  Processing: {fname} -> {treatment}")

            # Parse peaks
            all_peaks, filtered_peaks = parse_nist_txt(filepath)

            if not all_peaks:
                print(f"    WARNING: No peaks found in {fname}")
                continue

            # Enrich with NIST WebBook (only for filtered peaks with missing formulas)
            enriched_count = 0
            for p in filtered_peaks:
                has_formula = p.get('formula') and len(p.get('formula', '')) >= 2

                if enrich and not has_formula:
                    # Query WebBook
                    wb = enrich_compound_with_webbook(
                        p.get('name', ''),
                        p.get('cas', ''),
                        p.get('formula', '')
                    )
                    if wb:
                        if wb.get('formula'):
                            p['formula'] = wb['formula']
                            p['formula_source'] = 'NIST_WebBook'
                            # Recalculate atoms, NOSC, dG
                            p['atoms'] = parse_molecular_formula(wb['formula'])
                            nosc, dg = compute_nosc_and_dg(p['atoms'])
                            p['nosc'] = nosc
                            p['dg_cox'] = dg
                            enriched_count += 1
                        if wb.get('inchi'):
                            p['inchi'] = wb['inchi']
                        if wb.get('inchikey'):
                            p['inchikey'] = wb['inchikey']
                    time.sleep(enrich_delay)

            if enrich and enriched_count > 0:
                print(f"    Enriched {enriched_count} compounds from NIST WebBook")
                total_enriched += enriched_count

            # Classify each peak (all peaks and filtered)
            lookup = _load_lookup()
            for p in all_peaks:
                chen_group, chen_source = classify_compound_chen(p['name'], lookup)
                kal_group, kal_source = classify_compound_kallenbach(p['name'], lookup)
                p['chen_group'] = chen_group
                p['chen_source'] = chen_source
                p['kal_group'] = kal_group
                p['kal_source'] = kal_source

            # Classification already applied to all_peaks; filtered_peaks share the same dicts
            # (since they are the same objects) — no need to re-classify

            nist_total = sum(1 for p in all_peaks if p.get('si', 0) > 0)
            results[treatment] = {
                'sample_code': sample_code,
                'filepath': filepath,
                'mc_total_peaks': len(all_peaks),
                'nist_searched_peaks': nist_total,
                'si80_peaks': len(filtered_peaks),
                'all_peaks': all_peaks,
                'filtered_peaks': filtered_peaks
            }

            print(f"    MC={len(all_peaks)}, NIST_searched={nist_total}, SI>=80={len(filtered_peaks)}")

    if enrich:
        print(f"\n  Total enriched from NIST WebBook: {total_enriched} compounds")

    return results


def compute_statistics(results):
    """Compute source attribution and group statistics (on SI>=80 filtered data)."""
    stats = {}

    for treatment, data in results.items():
        peaks = data['filtered_peaks']
        total_area = sum(p['area'] for p in peaks) if peaks else 0

        # Chen statistics
        chen_source = defaultdict(float)
        chen_groups = defaultdict(float)

        # Kallenbach statistics
        kal_source = defaultdict(float)
        kal_groups = defaultdict(float)

        for p in peaks:
            area = p['area']
            area_pct = (area / total_area * 100) if total_area > 0 else 0

            chen_source[p['chen_source']] += area_pct
            chen_groups[p['chen_group']] += area_pct
            kal_source[p['kal_source']] += area_pct
            kal_groups[p['kal_group']] += area_pct

        stats[treatment] = {
            'total_area': total_area,
            'chen_source': dict(chen_source),
            'chen_groups': dict(chen_groups),
            'kal_source': dict(kal_source),
            'kal_groups': dict(kal_groups)
        }

    return stats


# ============================================================================
# 4. EXCEL OUTPUT
# ============================================================================

def write_excel(results, stats, output_path):
    """Write results to multi-sheet Excel workbook."""
    wb = Workbook()

    # Styles
    header_font = Font(name='Microsoft YaHei', bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    cell_font = Font(name='Microsoft YaHei', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # --- Sheet 1: 全化合物汇总 (ALL peaks, before SI filter) ---
    ws1 = wb.active
    ws1.title = '全化合物汇总'
    headers1 = ['处理', '峰号', '保留时间_min', '峰面积', 'SI', 'CAS',
                '化合物名称', '分子式', 'C', 'H', 'O', 'N', 'P', 'S', '其他元素',
                'H/C', 'O/C', '分子量',
                'Chen化学类别', 'Chen来源归属', 'Kallenbach化学类别', 'Kallenbach来源归属',
                'NOSC', 'ΔG_COX', 'SI>=80']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers1))

    row = 2
    for treatment, data in results.items():
        all_total_area = sum(p['area'] for p in data['all_peaks'])
        for p in data['all_peaks']:
            atoms = p.get('atoms', {})
            other_elem = ', '.join(f'{k}:{v}' for k, v in atoms.get('other', {}).items()) if atoms.get('other') else ''
            si = p.get('si', 0)
            passed = 'Yes' if si >= 80 else 'No'

            ws1.cell(row=row, column=1, value=treatment)
            ws1.cell(row=row, column=2, value=p['peak_num'])
            ws1.cell(row=row, column=3, value=round(p['ret_time'], 3))
            ws1.cell(row=row, column=4, value=p['area'])
            ws1.cell(row=row, column=5, value=si)
            ws1.cell(row=row, column=6, value=p.get('cas', ''))
            ws1.cell(row=row, column=7, value=p.get('name', ''))
            ws1.cell(row=row, column=8, value=p.get('formula', ''))
            ws1.cell(row=row, column=9, value=atoms.get('C', ''))
            ws1.cell(row=row, column=10, value=atoms.get('H', ''))
            ws1.cell(row=row, column=11, value=atoms.get('O', '') if atoms.get('O') else '')
            ws1.cell(row=row, column=12, value=atoms.get('N', '') if atoms.get('N') else '')
            ws1.cell(row=row, column=13, value=atoms.get('P', '') if atoms.get('P') else '')
            ws1.cell(row=row, column=14, value=atoms.get('S', '') if atoms.get('S') else '')
            ws1.cell(row=row, column=15, value=other_elem)
            ws1.cell(row=row, column=16, value=p.get('hc', ''))
            ws1.cell(row=row, column=17, value=p.get('oc', ''))
            ws1.cell(row=row, column=18, value=p.get('mol_weight', ''))
            ws1.cell(row=row, column=19, value=p.get('chen_group', ''))
            ws1.cell(row=row, column=20, value=p.get('chen_source', ''))
            ws1.cell(row=row, column=21, value=p.get('kal_group', ''))
            ws1.cell(row=row, column=22, value=p.get('kal_source', ''))
            ws1.cell(row=row, column=23, value=p.get('nosc', ''))
            ws1.cell(row=row, column=24, value=p.get('dg_cox', ''))
            ws1.cell(row=row, column=25, value=passed)
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=row, column=col).font = cell_font
                ws1.cell(row=row, column=col).border = thin_border
            row += 1

    # --- Sheet 2: 筛选统计 ---
    ws12 = wb.create_sheet('筛选统计')
    headers12 = ['样品代码', '处理', 'MC总峰数', 'NIST检索峰数', 'SI>=80峰数',
                 'SI>=80保留率_%', 'MC总面积', 'SI>=80总面积', 'SI>=80面积保留率_%',
                 'SI>=80唯一化合物数', '含卤素/其他元素化合物数']
    for col, h in enumerate(headers12, 1):
        ws12.cell(row=1, column=col, value=h)
    style_header(ws12, 1, len(headers12))

    row = 2
    for treatment, data in results.items():
        all_peaks = data['all_peaks']
        filtered = data['filtered_peaks']
        all_area = sum(p['area'] for p in all_peaks)
        si80_area = sum(p['area'] for p in filtered)
        unique = len(set(p['name'] for p in filtered if p.get('name')))
        exotic = sum(1 for p in filtered if p.get('atoms', {}).get('other'))

        ws12.cell(row=row, column=1, value=data['sample_code'])
        ws12.cell(row=row, column=2, value=treatment)
        ws12.cell(row=row, column=3, value=len(all_peaks))
        ws12.cell(row=row, column=4, value=data.get('nist_searched_peaks', 0))
        ws12.cell(row=row, column=5, value=len(filtered))
        ws12.cell(row=row, column=6, value=round(len(filtered)/max(len(all_peaks),1)*100, 1))
        ws12.cell(row=row, column=7, value=all_area)
        ws12.cell(row=row, column=8, value=si80_area)
        ws12.cell(row=row, column=9, value=round(si80_area/max(all_area,1)*100, 1))
        ws12.cell(row=row, column=10, value=unique)
        ws12.cell(row=row, column=11, value=exotic)
        for col in range(1, len(headers12) + 1):
            ws12.cell(row=row, column=col).font = cell_font
            ws12.cell(row=row, column=col).border = thin_border
        row += 1

    # --- Sheet 3: SI>=80 分类详细 (Chen) ---
    ws2 = wb.create_sheet('Chen_分类详细')
    headers2 = ['处理', '峰号', '保留时间_min', '峰面积', '相对丰度_%', 'SI', 'CAS',
                '化合物名称', '分子式', 'C', 'H', 'O', 'N', 'P', 'S', '其他元素',
                'H/C', 'O/C', '分子量', 'InChIKey',
                'Chen化学类别', 'Chen来源归属', 'NOSC', 'ΔG_COX']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))

    row = 2
    for treatment, data in results.items():
        total_area = sum(p['area'] for p in data['filtered_peaks'])
        for p in data['filtered_peaks']:
            ws2.cell(row=row, column=1, value=treatment)
            ws2.cell(row=row, column=2, value=p['peak_num'])
            ws2.cell(row=row, column=3, value=round(p['ret_time'], 3))
            ws2.cell(row=row, column=4, value=p['area'])
            ws2.cell(row=row, column=5, value=round(p['area']/total_area*100, 4) if total_area > 0 else 0)
            ws2.cell(row=row, column=6, value=p.get('si', ''))
            ws2.cell(row=row, column=7, value=p.get('cas', ''))
            atoms = p.get('atoms', {})
            other_elem = ', '.join(f'{k}:{v}' for k, v in atoms.get('other', {}).items()) if atoms.get('other') else ''

            ws2.cell(row=row, column=8, value=p['name'])
            ws2.cell(row=row, column=9, value=p.get('formula', ''))
            ws2.cell(row=row, column=10, value=atoms.get('C', ''))
            ws2.cell(row=row, column=11, value=atoms.get('H', ''))
            ws2.cell(row=row, column=12, value=atoms.get('O', '') if atoms.get('O') else '')
            ws2.cell(row=row, column=13, value=atoms.get('N', '') if atoms.get('N') else '')
            ws2.cell(row=row, column=14, value=atoms.get('P', '') if atoms.get('P') else '')
            ws2.cell(row=row, column=15, value=atoms.get('S', '') if atoms.get('S') else '')
            ws2.cell(row=row, column=16, value=other_elem)
            ws2.cell(row=row, column=17, value=p.get('hc', ''))
            ws2.cell(row=row, column=18, value=p.get('oc', ''))
            ws2.cell(row=row, column=19, value=p.get('mol_weight', ''))
            ws2.cell(row=row, column=20, value=p.get('inchikey', ''))
            ws2.cell(row=row, column=21, value=p['chen_group'])
            ws2.cell(row=row, column=22, value=p['chen_source'])
            ws2.cell(row=row, column=23, value=p.get('nosc', ''))
            ws2.cell(row=row, column=24, value=p.get('dg_cox', ''))
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row, column=col).font = cell_font
                ws2.cell(row=row, column=col).border = thin_border
            row += 1

    # --- Sheet 3: Chen 来源统计 ---
    ws3 = wb.create_sheet('Chen_来源统计')
    headers3 = ['处理', '来源类别', '峰数', '峰面积合计', '相对丰度_%']
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    row = 2
    for treatment, st in stats.items():
        for source in ['plant', 'microbial', 'mixed']:
            ws3.cell(row=row, column=1, value=treatment)
            ws3.cell(row=row, column=2, value=source)
            ws3.cell(row=row, column=3, value=sum(1 for p in results[treatment]['filtered_peaks'] if p['chen_source'] == source))
            area_sum = sum(p['area'] for p in results[treatment]['filtered_peaks'] if p['chen_source'] == source)
            ws3.cell(row=row, column=4, value=area_sum)
            ws3.cell(row=row, column=5, value=round(st['chen_source'].get(source, 0), 2))
            for col in range(1, len(headers3) + 1):
                ws3.cell(row=row, column=col).font = cell_font
                ws3.cell(row=row, column=col).border = thin_border
            row += 1

        # Add total row
        ws3.cell(row=row, column=1, value=treatment)
        ws3.cell(row=row, column=2, value='合计')
        ws3.cell(row=row, column=3, value=len(results[treatment]['filtered_peaks']))
        ws3.cell(row=row, column=4, value=st['total_area'])
        ws3.cell(row=row, column=5, value=100.0)
        for col in range(1, len(headers3) + 1):
            ws3.cell(row=row, column=col).font = Font(name='Microsoft YaHei', bold=True, size=10)
            ws3.cell(row=row, column=col).border = thin_border
        row += 1

    # --- Sheet 4: Chen 来源饼图数据 ---
    ws4 = wb.create_sheet('Chen_来源饼图')
    headers4 = ['处理', '植物源', '微生物源', '混合来源', '合计']
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers4))

    row = 2
    for treatment, st in stats.items():
        ws4.cell(row=row, column=1, value=treatment)
        ws4.cell(row=row, column=2, value=round(st['chen_source'].get('plant', 0), 4))
        ws4.cell(row=row, column=3, value=round(st['chen_source'].get('microbial', 0), 4))
        ws4.cell(row=row, column=4, value=round(st['chen_source'].get('mixed', 0), 4))
        ws4.cell(row=row, column=5, value=100.0)
        for col in range(1, len(headers4) + 1):
            ws4.cell(row=row, column=col).font = cell_font
            ws4.cell(row=row, column=col).border = thin_border
        row += 1

    # --- Sheet 5: Kallenbach 分类详细 ---
    ws5 = wb.create_sheet('Kallenbach_分类详细')
    headers5 = ['处理', '峰号', '保留时间_min', '峰面积', '相对丰度_%', 'SI', 'CAS',
                '化合物名称', '分子式', 'C', 'H', 'O', 'N', 'P', 'S', '其他元素',
                'H/C', 'O/C', '分子量', 'InChIKey',
                'Kallenbach化学类别', 'Kallenbach来源归属']
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    style_header(ws5, 1, len(headers5))

    row = 2
    for treatment, data in results.items():
        total_area = sum(p['area'] for p in data['filtered_peaks'])
        for p in data['filtered_peaks']:
            ws5.cell(row=row, column=1, value=treatment)
            ws5.cell(row=row, column=2, value=p['peak_num'])
            ws5.cell(row=row, column=3, value=round(p['ret_time'], 3))
            ws5.cell(row=row, column=4, value=p['area'])
            ws5.cell(row=row, column=5, value=round(p['area']/total_area*100, 4) if total_area > 0 else 0)
            ws5.cell(row=row, column=6, value=p.get('si', ''))
            ws5.cell(row=row, column=7, value=p.get('cas', ''))
            atoms = p.get('atoms', {})
            other_elem = ', '.join(f'{k}:{v}' for k, v in atoms.get('other', {}).items()) if atoms.get('other') else ''

            ws5.cell(row=row, column=8, value=p['name'])
            ws5.cell(row=row, column=9, value=p.get('formula', ''))
            ws5.cell(row=row, column=10, value=atoms.get('C', ''))
            ws5.cell(row=row, column=11, value=atoms.get('H', ''))
            ws5.cell(row=row, column=12, value=atoms.get('O', '') if atoms.get('O') else '')
            ws5.cell(row=row, column=13, value=atoms.get('N', '') if atoms.get('N') else '')
            ws5.cell(row=row, column=14, value=atoms.get('P', '') if atoms.get('P') else '')
            ws5.cell(row=row, column=15, value=atoms.get('S', '') if atoms.get('S') else '')
            ws5.cell(row=row, column=16, value=other_elem)
            ws5.cell(row=row, column=17, value=p.get('hc', ''))
            ws5.cell(row=row, column=18, value=p.get('oc', ''))
            ws5.cell(row=row, column=19, value=p.get('mol_weight', ''))
            ws5.cell(row=row, column=20, value=p.get('inchikey', ''))
            ws5.cell(row=row, column=21, value=p['kal_group'])
            ws5.cell(row=row, column=22, value=p['kal_source'])
            for col in range(1, len(headers5) + 1):
                ws5.cell(row=row, column=col).font = cell_font
                ws5.cell(row=row, column=col).border = thin_border
            row += 1

    # --- Sheet 6: Kallenbach 来源统计 ---
    ws6 = wb.create_sheet('Kallenbach_来源统计')
    headers6 = ['处理', '来源类别', '峰数', '峰面积合计', '相对丰度_%']
    for col, h in enumerate(headers6, 1):
        ws6.cell(row=1, column=col, value=h)
    style_header(ws6, 1, len(headers6))

    row = 2
    for treatment, st in stats.items():
        for source in ['plant', 'microbial', 'mixed']:
            ws6.cell(row=row, column=1, value=treatment)
            ws6.cell(row=row, column=2, value=source)
            count = sum(1 for p in results[treatment]['filtered_peaks'] if p['kal_source'] == source)
            area_sum = sum(p['area'] for p in results[treatment]['filtered_peaks'] if p['kal_source'] == source)
            ws6.cell(row=row, column=3, value=count)
            ws6.cell(row=row, column=4, value=area_sum)
            ws6.cell(row=row, column=5, value=round(st['kal_source'].get(source, 0), 2))
            for col in range(1, len(headers6) + 1):
                ws6.cell(row=row, column=col).font = cell_font
                ws6.cell(row=row, column=col).border = thin_border
            row += 1

        ws6.cell(row=row, column=1, value=treatment)
        ws6.cell(row=row, column=2, value='合计')
        ws6.cell(row=row, column=3, value=len(results[treatment]['filtered_peaks']))
        ws6.cell(row=row, column=4, value=st['total_area'])
        ws6.cell(row=row, column=5, value=100.0)
        for col in range(1, len(headers6) + 1):
            ws6.cell(row=row, column=col).font = Font(name='Microsoft YaHei', bold=True, size=10)
            ws6.cell(row=row, column=col).border = thin_border
        row += 1

    # --- Sheet 7: Kallenbach 饼图 ---
    ws7 = wb.create_sheet('Kallenbach_来源饼图')
    headers7 = ['处理', '植物源', '微生物源', '混合来源', '合计']
    for col, h in enumerate(headers7, 1):
        ws7.cell(row=1, column=col, value=h)
    style_header(ws7, 1, len(headers7))

    row = 2
    for treatment, st in stats.items():
        ws7.cell(row=row, column=1, value=treatment)
        ws7.cell(row=row, column=2, value=round(st['kal_source'].get('plant', 0), 4))
        ws7.cell(row=row, column=3, value=round(st['kal_source'].get('microbial', 0), 4))
        ws7.cell(row=row, column=4, value=round(st['kal_source'].get('mixed', 0), 4))
        ws7.cell(row=row, column=5, value=100.0)
        for col in range(1, len(headers7) + 1):
            ws7.cell(row=row, column=col).font = cell_font
            ws7.cell(row=row, column=col).border = thin_border
        row += 1

    # --- Sheet 8: 元素比统计 (H/C, O/C by chemical class) ---
    ws8 = wb.create_sheet('元素比统计_Chen')
    headers8 = ['处理', '化学类别', '化合物数', '峰面积合计', '面积加权H/C', '面积加权O/C',
                '算术平均H/C', '算术平均O/C', 'H/C最小值', 'H/C最大值', 'O/C最小值', 'O/C最大值']
    for col, h in enumerate(headers8, 1):
        ws8.cell(row=1, column=col, value=h)
    style_header(ws8, 1, len(headers8))

    row = 2
    for treatment, data in results.items():
        peaks = data['filtered_peaks']
        class_data = defaultdict(list)
        for p in peaks:
            if p.get('hc') is not None and p.get('oc') is not None:
                class_data[p['chen_group']].append(p)

        for chen_class in sorted(class_data.keys()):
            cpeaks = class_data[chen_class]
            total_area = sum(p['area'] for p in cpeaks)
            if total_area == 0:
                continue
            # Area-weighted H/C and O/C
            w_hc = sum(p['hc'] * p['area'] for p in cpeaks) / total_area
            w_oc = sum(p['oc'] * p['area'] for p in cpeaks) / total_area
            # Arithmetic mean
            a_hc = sum(p['hc'] for p in cpeaks) / len(cpeaks)
            a_oc = sum(p['oc'] for p in cpeaks) / len(cpeaks)
            hc_vals = [p['hc'] for p in cpeaks]
            oc_vals = [p['oc'] for p in cpeaks]

            ws8.cell(row=row, column=1, value=treatment)
            ws8.cell(row=row, column=2, value=chen_class)
            ws8.cell(row=row, column=3, value=len(cpeaks))
            ws8.cell(row=row, column=4, value=total_area)
            ws8.cell(row=row, column=5, value=round(w_hc, 4))
            ws8.cell(row=row, column=6, value=round(w_oc, 4))
            ws8.cell(row=row, column=7, value=round(a_hc, 4))
            ws8.cell(row=row, column=8, value=round(a_oc, 4))
            ws8.cell(row=row, column=9, value=round(min(hc_vals), 4))
            ws8.cell(row=row, column=10, value=round(max(hc_vals), 4))
            ws8.cell(row=row, column=11, value=round(min(oc_vals), 4))
            ws8.cell(row=row, column=12, value=round(max(oc_vals), 4))
            for col in range(1, len(headers8) + 1):
                ws8.cell(row=row, column=col).font = cell_font
                ws8.cell(row=row, column=col).border = thin_border
            row += 1

    # --- Sheet 9: 元素比统计_Kallenbach ---
    ws9 = wb.create_sheet('元素比统计_Kallenbach')
    headers9 = ['处理', '化学类别', '化合物数', '峰面积合计', '面积加权H/C', '面积加权O/C',
                '算术平均H/C', '算术平均O/C', 'H/C最小值', 'H/C最大值', 'O/C最小值', 'O/C最大值']
    for col, h in enumerate(headers9, 1):
        ws9.cell(row=1, column=col, value=h)
    style_header(ws9, 1, len(headers9))

    row = 2
    for treatment, data in results.items():
        peaks = data['filtered_peaks']
        class_data = defaultdict(list)
        for p in peaks:
            if p.get('hc') is not None and p.get('oc') is not None:
                class_data[p['kal_group']].append(p)

        for kal_class in sorted(class_data.keys()):
            cpeaks = class_data[kal_class]
            total_area = sum(p['area'] for p in cpeaks)
            if total_area == 0:
                continue
            w_hc = sum(p['hc'] * p['area'] for p in cpeaks) / total_area
            w_oc = sum(p['oc'] * p['area'] for p in cpeaks) / total_area
            a_hc = sum(p['hc'] for p in cpeaks) / len(cpeaks)
            a_oc = sum(p['oc'] for p in cpeaks) / len(cpeaks)
            hc_vals = [p['hc'] for p in cpeaks]
            oc_vals = [p['oc'] for p in cpeaks]

            ws9.cell(row=row, column=1, value=treatment)
            ws9.cell(row=row, column=2, value=kal_class)
            ws9.cell(row=row, column=3, value=len(cpeaks))
            ws9.cell(row=row, column=4, value=total_area)
            ws9.cell(row=row, column=5, value=round(w_hc, 4))
            ws9.cell(row=row, column=6, value=round(w_oc, 4))
            ws9.cell(row=row, column=7, value=round(a_hc, 4))
            ws9.cell(row=row, column=8, value=round(a_oc, 4))
            ws9.cell(row=row, column=9, value=round(min(hc_vals), 4))
            ws9.cell(row=row, column=10, value=round(max(hc_vals), 4))
            ws9.cell(row=row, column=11, value=round(min(oc_vals), 4))
            ws9.cell(row=row, column=12, value=round(max(oc_vals), 4))
            for col in range(1, len(headers9) + 1):
                ws9.cell(row=row, column=col).font = cell_font
                ws9.cell(row=row, column=col).border = thin_border
            row += 1

    # --- Sheet 10: 说明_QC ---
    ws8 = wb.create_sheet('说明_QC')
    info = [
        ['Py-GC-MS 批量分析报告'],
        [''],
        ['项目', '说明'],
        ['1. 分析方法', '基于NIST搜索导出TXT文件的MC Peak Table解析'],
        ['2. SI筛选', '保留SI>=80的化合物（对应NIST Match Factor >= 80%）'],
        ['3. Chen 2023分类', '参考Chen et al. 2023, Carbon Research 2:1. DOI: 10.1007/s44246-022-00034-0'],
        ['4. Kallenbach 2016分类', '参考Kallenbach et al. 2016, Nature Communications 7:13630. DOI: 10.1038/ncomms13630'],
        ['5. 来源归属', '植物源(plant)、微生物源(microbial)、混合来源(mixed)'],
        ['6. 化学类别(Chen)', 'lipids, monocyclic_aromatics, polycyclic_aromatics, phenolics, polysaccharides, lignins, amino_N_bearing, heterocyclic_N_bearing, other_N_bearing, unspecified'],
        ['7. 化学类别(Kallenbach)', 'lipids, lignin_derivatives, polysaccharides, proteins, non_protein_N, phenolics, aromatics, unspecified'],
        ['8. NIST WebBook', '通过 --enrich 启用在线查询补全缺失的分子式/InChIKey'],
        ['9. 注意', '分类基于286条精确映射表 + 关键词兜底。对于需验证的化合物请人工核对。'],
        ['10. 脚本', 'pygcms-batch skill v1.1']
    ]
    for i, row_data in enumerate(info, 1):
        for j, val in enumerate(row_data, 1):
            ws8.cell(row=i, column=j, value=val)
            if i == 1:
                ws8.cell(row=i, column=j).font = Font(name='Microsoft YaHei', bold=True, size=14)
            elif i == 3:
                ws8.cell(row=i, column=j).font = header_font
            elif i > 3:
                ws8.cell(row=i, column=j).font = cell_font

    # Adjust column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[col_letter].width = adjusted_width

    # Save
    wb.save(output_path)
    print(f"\nOutput saved to: {output_path}")


# ============================================================================
# 5. MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Py-GC-MS Batch Analysis - Compound classification by Chen 2023 and Kallenbach 2016'
    )
    parser.add_argument('--input', '-i', required=True,
                        help='Input directory containing NIST TXT export files')
    parser.add_argument('--output', '-o', required=True,
                        help='Output Excel file path (.xlsx)')
    parser.add_argument('--sample_map', '-m', default=None,
                        help='JSON file mapping sample codes to treatment names')
    parser.add_argument('--enrich', '-e', action='store_true',
                        help='Query NIST Chemistry WebBook to fill missing molecular formulas')
    parser.add_argument('--enrich_delay', type=float, default=0.5,
                        help='Delay in seconds between WebBook queries (default: 0.5)')

    args = parser.parse_args()

    # Load sample mapping
    sample_map = {}
    if args.sample_map and os.path.exists(args.sample_map):
        with open(args.sample_map, 'r', encoding='utf-8') as f:
            sample_map = json.load(f)

    print("=" * 60)
    print("Py-GC-MS Batch Analysis")
    print("=" * 60)
    print(f"Input directory: {args.input}")
    print(f"Output file: {args.output}")
    print(f"Sample mapping: {sample_map if sample_map else 'None (using filenames)'}")
    print()

    # Process
    results = process_directory(args.input, sample_map, enrich=args.enrich,
                                enrich_delay=args.enrich_delay)

    if not results:
        print("ERROR: No valid TXT files found or no peaks extracted!")
        return

    # Statistics
    stats = compute_statistics(results)

    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY: Chen 2023 Source Attribution")
    print("=" * 60)
    for treatment, st in stats.items():
        print(f"\n{treatment}:")
        d = results[treatment]
        print(f"  MC peaks: {d.get('mc_total_peaks','?')} | NIST searched: {d.get('nist_searched_peaks','?')} | SI>=80: {d.get('si80_peaks','?')}")
        print(f"  Total area (SI>=80): {st['total_area']:,.0f}")
        for source in ['plant', 'microbial', 'mixed']:
            val = st['chen_source'].get(source, 0)
            print(f"  {source}: {val:.2f}%")

    # Write Excel
    write_excel(results, stats, args.output)

    print("\nDone!")


if __name__ == '__main__':
    main()
